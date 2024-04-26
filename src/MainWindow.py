from math import ceil

from PyQt5.QtWidgets import QLabel, QLineEdit, QApplication, QMainWindow, QPushButton, QWidget, QVBoxLayout, \
    QTableWidget, QComboBox, QStackedWidget,QAction,QHeaderView,QTableWidgetItem
from PyQt5.QtCore import QSize
from PyQt5.QtGui import QIcon
from PyQt5 import QtWidgets
from openpyxl.chart import BarChart, Reference, PieChart, SurfaceChart, PieChart3D

from start import Ui_MainWindowSTART
from add_bpla import Ui_MainWindowADD
from add_polet import Ui_MainWindowADD_POLET
from analitika_po_effect import Ui_MainWindow_Effect
from analitika_po_jivuchesti import Ui_MainWindow_Jivuchest
from analitika_po_resursozatratnosti import Ui_MainWindow_Resurs
from view_Polet import Ui_MainWindow_Polet
from edit_bpla import Ui_MainWindow_Edit
from view_analit_po_jivuchesti import Ui_MainWindow_Po_Jivuchesti
from view_analit_po_rashod_topliva import Ui_MainWindow_PO_Topliva
from view_analit_po_effect import Ui_MainWindow_Po_Effect
from view_BPLA import Ui_MainWindow_Show_BPLA
from Bpla import Bpla
from Polet import Polet
import psycopg2 as pg
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.styles import (Border, Side, Alignment, Font)
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl import chart
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import RichTextProperties


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.cursor = ""
        self.connect = ""
        self.index_to_clicked_bpla_and_delete=""
        self.idndex_to_clicked_flight_adn_delete=""
        self.vivod_v_fail_ef_bpla = ""
        self.vivod_v_fail_jiv_bpla = ""
        self.vivod_v_fail_res_bpla = ""
        self.index_all_models = []
        self.index_effective=""
        self.index_jivuchesti=""
        self.index_resursozatratnosti=""
        self.all_models=[]
        self.connect_to_bd()
        self.setWindowTitle("Тур 3")
        self.setFixedSize(400,600)
        self.first_widget = QMainWindow()
        self.ui_first = Ui_MainWindowSTART()
        self.ui_first.setupUi(self.first_widget)
        self.ui_first.pushButton_ADD_BPLA.clicked.connect(self.show_ADD_BPLA_window)
        self.ui_first.pushButton_ADD_FLIGHTS.clicked.connect(self.show_ADD_POLET_window)
        self.ui_first.pushButton_EFFEKTIVNOST.clicked.connect(self.show_Analitika_Effectivnosti)
        self.ui_first.pushButton_JIVUCHEST.clicked.connect(self.show_Analitika_Jivuchesti)
        self.ui_first.pushButton_RESURSOZATRATNOST.clicked.connect(self.show_Analitika_Resursozatratnosti)
        self.ui_first.pushButton_ViVESTI_BPLA.clicked.connect(self.show_BPLA)
        self.ui_first.pushButton_VIVESTI_FLIGHTS.clicked.connect(self.show_Polet)
        self.ui_first.pushButton_CHANGE_BPLA.clicked.connect(self.change_bpla)
        self.ui_first.pushButton_DELETE_BPLA.clicked.connect(self.delete_bpla)
        self.ui_first.pushButton_DELETE_FLIGHTS.clicked.connect(self.delete_flight)

        self.second_widget = QMainWindow()
        self.ui_second = Ui_MainWindowADD()
        self.ui_second.setupUi(self.second_widget)
        self.ui_second.pushButton_ADD_BPLA_IN_BD.clicked.connect(self.add_BPLA_InBd)
        self.action = QtWidgets.QAction("Назад" , self.second_widget)
        self.action.triggered.connect(self.show_start_window)
        self.ui_second.menu.addAction(self.action)
        self.ui_second.menu.setStyleSheet("""
            QMenu {
                background-color: rgb(245, 245, 245); /* sets background of the menu */
                color: black; /* sets color of menu text */
            }
            QMenu::item:selected {
                color: rgb(0, 120, 215);
                background-color: rgb(245, 245, 245);
            }
        """)

        self.effect_widget = QMainWindow()
        self.ui_an_effe = Ui_MainWindow_Po_Effect()
        self.ui_an_effe.setupUi(self.effect_widget)
        self.act_po_eff = QtWidgets.QAction("Вывести в отчет",self.effect_widget)
        self.act_po_eff.triggered.connect(self.vivod_v_fail_effect)
        self.ui_an_effe.menu.addAction(self.act_po_eff)
        self.ui_an_effe.menu.setStyleSheet('''
            QMenu::item:selected{
                    color:rgb(0, 120, 215);
                    background-color:rgb(245, 245, 245);
            }
        ''')
        self.ui_an_effe.menu.setEnabled(False)
        self.ui_an_effe.tableWidget.selectionModel().selectionChanged.connect(self.can_vivod)


        self.an_res = QMainWindow()
        self.ui_an_res = Ui_MainWindow_PO_Topliva()
        self.ui_an_res.setupUi(self.an_res)
        self.act_po_res = QtWidgets.QAction("Вывести в отчет", self.an_res)
        self.act_po_res.triggered.connect(self.vivod_v_fail_res)
        self.ui_an_res.menu.addAction(self.act_po_res)
        self.ui_an_res.menu.setEnabled(False)
        self.ui_an_res.menu.setStyleSheet('''
                    QMenu::item:selected{
                            color:rgb(0, 120, 215);
                            background-color:rgb(245, 245, 245);
                    }
                ''')
        self.ui_an_res.tableWidget.selectionModel().selectionChanged.connect(self.can_vivod3)

        self.an_jiv = QMainWindow()
        self.ui_an_jiv = Ui_MainWindow_Po_Jivuchesti()
        self.ui_an_jiv.setupUi(self.an_jiv)
        self.act_po_jiv = QtWidgets.QAction("Вывести в отчет", self.an_jiv)
        self.act_po_jiv.triggered.connect(self.vivod_v_fail_jiv)
        self.ui_an_jiv.menu.addAction(self.act_po_jiv)
        self.ui_an_jiv.menu.setEnabled(False)
        self.ui_an_jiv.menu.setStyleSheet('''
                            QMenu::item:selected{
                                    color:rgb(0, 120, 215);
                                    background-color:rgb(245, 245, 245);
                            }
                        ''')
        self.ui_an_jiv.tableWidget.selectionModel().selectionChanged.connect(self.can_vivo2)


        self.third_widget = QMainWindow()
        self.ui_third = Ui_MainWindowADD_POLET()
        self.ui_third.setupUi(self.third_widget)
        self.ui_third.pushButton_ADD_POLET_IN_BD.clicked.connect(self.addPoletInBd)
        self.action_add_polet = QtWidgets.QAction("Назад",self.third_widget)
        self.action_add_polet.triggered.connect(self.show_start_window)
        self.ui_third.menu.addAction(self.action_add_polet)
        self.ui_third.menu.setStyleSheet("""
            QMenu::item:selected{
                    color:rgb(0, 120, 215);
                    background-color:rgb(245, 245, 245);
            }
        """)

        self.fourth_widget = QMainWindow()
        self.ui_fourth = Ui_MainWindow_Effect()
        self.ui_fourth.setupUi(self.fourth_widget)
        self.ui_fourth.pushButton_VIVESTI_ANALITIKU_IS_BD.clicked.connect(self.analit_effect)

        self.fives_widget = QMainWindow()
        self.ui_fives = Ui_MainWindow_Jivuchest()
        self.ui_fives.setupUi(self.fives_widget)
        self.ui_fives.pushButton_VIVESTI_ANALITIKU_IS_BD.clicked.connect(self.analit_jivuch)

        self.six_widget = QMainWindow()
        self.ui_six = Ui_MainWindow_Resurs()
        self.ui_six.setupUi(self.six_widget)
        self.ui_six.pushButton_VIVESTI_ANALITIKU_IS_BD.clicked.connect(self.analit_resurs)

        self.seven_widget = QMainWindow()
        self.ui_seven = Ui_MainWindow_Show_BPLA()
        self.ui_seven.setupUi(self.seven_widget)
        self.ui_seven.tableWidget.selectionModel().selectionChanged.connect(self.enable_delete_edit)

        self.eight_widget = QMainWindow()
        self.ui_eight = Ui_MainWindow_Polet()
        self.ui_eight.setupUi(self.eight_widget)
        self.ui_eight.tableWidget.selectionModel().selectionChanged.connect(self.selectionChanged)

        self.nine_widget = QMainWindow()
        self.ui_nine = Ui_MainWindow_Edit()
        self.ui_nine.setupUi(self.nine_widget)
        self.ui_nine.pushButton_EDIT_BPLA_IN_BD.clicked.connect(self.edit_bpla_in_bd)
        self.act_edit = QtWidgets.QAction("Назад",self.nine_widget)
        self.act_edit.triggered.connect(self.show_start_window)
        self.ui_nine.menu.addAction(self.act_edit)
        self.ui_nine.menu.setStyleSheet("""
            QMenu::item:selected{
                color:rgb(0, 120, 215);
                background-color:rgb(245, 245, 245);
            }
        """)


        self.stacked_widget = QtWidgets.QStackedWidget()
        self.stacked_widget.addWidget(self.first_widget)
        self.stacked_widget.addWidget(self.second_widget)
        self.stacked_widget.addWidget(self.third_widget)
        self.stacked_widget.addWidget(self.nine_widget)
        self.setCentralWidget(self.stacked_widget)

    def show_start_window(self):
        self.setWindowTitle("Тур 3")
        self.stacked_widget.setCurrentIndex(0)#Это индекс старт окна
        self.setFixedSize(400, 600)

    def show_ADD_BPLA_window(self):
        self.setWindowTitle("Добавить БПЛА")
        self.stacked_widget.setCurrentIndex(1)
        self.setFixedSize(500,430)

    def add_BPLA_InBd(self):
        mas = []
        try:
            mas.append(self.ui_second.lineEdit_NAIMENOVANIE_BPLA.text())
            mas.append(self.ui_second.lineEdit_MODEL_BPLA.text())
            mas.append(int(self.ui_second.lineEdit_ZAVODSKOI_NOMER_BPLA.text()))
            mas.append(float(self.ui_second.lineEdit_VES_BPLA.text()))
            mas.append(float(self.ui_second.lineEdit_MAX_SPEED_BPLA.text()))
            mas.append(float(self.ui_second.lineEdit_MAX_FLIGHTS_TIME_BPLA.text()))
            mas.append(float(self.ui_second.lineEdit_MAX_FLIGHTS_DIST_BPLA.text()))
            bpla = Bpla(None, None, None, None, None, None, None, None)
            try:
                bpla.add_data(mas, self.connect, self.cursor)
            except ValueError as e:
                print(f"Error: {e}")
            except Exception as err:
                print(f"Error: {err}")
                self.connect_to_bd()
        except Exception as err:
            print(f"Error: {err}")


    def show_ADD_POLET_window(self):
        self.setWindowTitle("Добавить Полет")
        self.stacked_widget.setCurrentIndex(2)
        self.setFixedSize(500,430)

    def addPoletInBd(self):
        mas = []
        mas.append(self.ui_third.lineEdit_ID_BPLA.text())
        mas.append(self.ui_third.lineEdit__FLIGHT_TIME_POLET.text())
        mas.append(self.ui_third.lineEdit_DIST_POLET.text())
        mas.append(self.ui_third.lineEdit_MAX_HEIGHT_FLIGHT.text())
        mas.append(self.ui_third.lineEdit_BOEVIE_TASKS_COMPLETE.text())
        vived_is_stroi = self.ui_third.checkBox_VIVEDEN_IS_STOROIA.isChecked()
        mas.append(self.ui_third.lineEdit_ZATRACHENO_TOPLIVO.text())
        try:
            polet = Polet(None, None, None, None, None, None, None, None)
            polet.add(mas,vived_is_stroi,self.cursor,self.connect)
        except Exception as err:
            print(f"Errro: {err}")
            self.connect_to_bd()

    def show_Analitika_Effectivnosti(self):
        bpla = Bpla(None, None, None, None, None, None, None, None)
        data = bpla.list(self.cursor)
        for i, row in enumerate(data):
            model = row[2]
            self.ui_fourth.comboBox_VIBOR_MODEL.addItem("")
            self.ui_fourth.comboBox_VIBOR_MODEL.setItemText(i+1,model)

        self.fourth_widget.show()

    def show_Analitika_Jivuchesti(self):
        bpla = Bpla(None, None,
                    None, None, None, None, None, None)
        data = bpla.list(self.cursor)
        for i,row in enumerate(data):
            model = row[2]
            self.ui_fives.comboBox_VIBOR_MODEL.addItem("")
            self.ui_fives.comboBox_VIBOR_MODEL.setItemText(i + 1, model)
        self.fives_widget.show()

    def show_Analitika_Resursozatratnosti(self):
        bpla = Bpla(None, None, None, None, None, None, None, None)
        data = bpla.list(self.cursor)
        for i, row in enumerate(data):
            model = row[2]
            self.ui_six.comboBox_VIBOR_MODEL.addItem("")
            self.ui_six.comboBox_VIBOR_MODEL.setItemText(i + 1, model)
        self.six_widget.show()

    def show_BPLA(self):
        bpla = Bpla(None, None, None, None, None, None, None, None)
        try:
            data = bpla.list(self.cursor)
            self.ui_seven.tableWidget.setRowCount(len(data))
            for i, row in enumerate(data):
                for j, cell in enumerate(row):
                    item = QtWidgets.QTableWidgetItem()
                    item.setText(str(cell))  # Устанавливаем текст элемента как данные из ячейки
                    self.ui_seven.tableWidget.setItem(i, j, item)  # Добавляем элемент в таблицу


        except Exception as err:
            print(f"Error: {err}")
            self.connect_to_bd()

        headers = [""]*self.ui_seven.tableWidget.rowCount()
        self.ui_seven.tableWidget.setVerticalHeaderLabels(headers)
        self.seven_widget.show()


    def show_Polet(self):
        polet = Polet(None, None, None, None, None, None, None, None)
        try:
            data = polet.list(self.cursor)
            self.ui_eight.tableWidget.setRowCount(len(data))
            for i,row in enumerate(data):
                for j,cell in enumerate(row):
                    item = QtWidgets.QTableWidgetItem()
                    item.setText(str(cell))
                    self.ui_eight.tableWidget.setItem(i,j,item)
        except Exception as err:
            print(f"Error: {err}")
            self.connect_to_bd()

        headers = [""] * self.ui_eight.tableWidget.rowCount()
        self.ui_eight.tableWidget.setVerticalHeaderLabels(headers)
        self.eight_widget.show()

    def edit_bpla_in_bd(self):
        mas1 = []
        print(type(self.ui_nine.lineEdit_NAME.text()))
        print(type(self.ui_nine.lineEdit_MODEL.text()))
        mas1.append(self.ui_nine.lineEdit_NAME.text())
        mas1.append(self.ui_nine.lineEdit_MODEL.text())
        mas1.append(self.ui_nine.lineEdit_ZAVODCKOY_NOMER.text())
        mas1.append(self.ui_nine.lineEdit_VES.text())
        mas1.append(self.ui_nine.lineEdit_MAX_SPEED.text())
        mas1.append(self.ui_nine.lineEdit_MAX_TIME_FLIGHT.text())
        mas1.append(self.ui_nine.lineEdit_MAX_DIST_POLET.text())

        try:
            bpla = Bpla(None, None, None, None, None, None, None, None)
            bpla.edit(mas1,self.index_to_clicked_bpla_and_delete,self.connect,self.cursor)
        except Exception as err:
            print(f"Error: {err}")
            self.connect_to_bd()

    def delete_bpla(self):
        try:
            bpla = Bpla(None, None, None, None, None, None, None, None)
            bpla.delete(self.index_to_clicked_bpla_and_delete,self.connect,self.cursor)
        except Exception as err:
            print(f"Error: {err}")
            self.connect_to_bd()

    def delete_flight(self):
        try:
            polet = polet = Polet(None, None, None, None, None, None, None, None)
            polet.delete(self.idndex_to_clicked_flight_adn_delete,self.cursor,self.connect)
        except Exception as err:
            print(f"Error:{err}")
            self.connect_to_bd()

    def enable_delete_edit(self):
        stroka = self.ui_seven.tableWidget.selectionModel().selectedRows()
        if stroka:
            row = stroka[0].row()
            self.index_to_clicked_bpla_and_delete = self.ui_seven.tableWidget.item(row,0).text()
            self.ui_first.pushButton_DELETE_BPLA.setEnabled(True)
            self.ui_first.pushButton_CHANGE_BPLA.setEnabled(True)
        else:
            self.ui_first.pushButton_DELETE_BPLA.setEnabled(False)
            self.ui_first.pushButton_CHANGE_BPLA.setEnabled(False)


    def selectionChanged(self,selected,deselected):
        selected_rows = self.ui_eight.tableWidget.selectionModel().selectedRows()
        if selected_rows:
            row = selected_rows[0].row()
            self.idndex_to_clicked_flight_adn_delete = self.ui_eight.tableWidget.item(row,0).text()
            self.ui_first.pushButton_DELETE_FLIGHTS.setEnabled(True)
        else:
            self.ui_first.pushButton_DELETE_FLIGHTS.setEnabled(False)

    def change_bpla(self):
        self.setWindowTitle("Редактировать БПЛА")
        self.stacked_widget.setCurrentIndex(3)
        self.setFixedSize(500, 430)

    def analit_resurs(self):
        sort_po = self.ui_six.comboBox_SORT_PO.currentText()
        model = self.ui_six.comboBox_VIBOR_MODEL.currentText()
        if model != "Не выбрано":
            self.all_models=[]
            self.cursor.execute('''SELECT * FROM public.bpla WHERE model_bpla=%s''', (model,))
            self.all_models.append(self.cursor.fetchone())
            self.cursor.execute('''SELECT * FROM public.polet WHERE uniq_id_bpla=%s''', (self.all_models[0][0],))
            flight_sam = self.cursor.fetchall()
            toplivo_polet = 0
            for row in flight_sam:
                toplivo_polet = toplivo_polet + row[7]
            #НЕЛЬЗЯ ДЕЛИТЬ НА НОЛЬ
            #ЕСЛИ МОДЕЛЬ НЕ ВЫБРАНА,ТО ВЫВОДИМ ДЛЯ ВСЕХ И СОРТИРУЕМ
            #И ЛУЧШЕ НА ВСЯКИЙ,ДАЖЕ ЕСЛИ ВЫБРАНА,ВСЕ РАВНО СОРТИРУЕМ
            self.index_resursozatratnosti = toplivo_polet/len(flight_sam)
            headers = [""] * self.ui_an_res.tableWidget.rowCount()
            self.ui_an_res.tableWidget.setVerticalHeaderLabels(headers)

            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][0]))
            self.ui_an_res.tableWidget.setItem(0, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][1]))
            self.ui_an_res.tableWidget.setItem(0, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][2]))
            self.ui_an_res.tableWidget.setItem(0, 2, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.index_resursozatratnosti))
            self.ui_an_res.tableWidget.setItem(0, 3, item)
            self.an_res.show()
        else:
            self.all_models=[]
            self.cursor.execute('''SELECT * FROM public.bpla''')
            self.all_models = self.cursor.fetchall()
            self.index_all_models = []
            for row in self.all_models:
                self.cursor.execute('''SELECT * FROM public.polet WHERE uniq_id_bpla=%s''',(row[0],))
                data = self.cursor.fetchall()
                toplivo_polet = 0
                for item in data:
                    toplivo_polet = toplivo_polet + item[7]
                self.index_all_models.append(toplivo_polet/len(data))
            if sort_po == "По возрастанию":
                self.index_all_models.sort()
            elif sort_po == "По убыванию":
                self.index_all_models.sort(reverse=True)

            self.ui_an_res.tableWidget.setRowCount(len(self.index_all_models))
            headers = [""]*self.ui_an_res.tableWidget.rowCount()
            self.ui_an_res.tableWidget.setVerticalHeaderLabels(headers)

            for i,index in enumerate(self.index_all_models):
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][0]))
                self.ui_an_res.tableWidget.setItem(i,0,item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][1]))
                self.ui_an_res.tableWidget.setItem(i, 1, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][2]))
                self.ui_an_res.tableWidget.setItem(i, 2, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.index_all_models[i]))
                self.ui_an_res.tableWidget.setItem(i, 3, item)

        self.an_res.show()

    def analit_jivuch(self):
        sort_po = self.ui_fives.comboBox_SORT_PO.currentText()
        model = self.ui_fives.comboBox_VIBOR_MODEL.currentText()
        # сначало считать по убыванию ил по возрастанию
        # затем считать по какой модели и вывести
        if model != "Не выбрано":
            self.all_models=[]
            self.cursor.execute('''SELECT * FROM public.bpla WHERE model_bpla=%s''', (model,))
            self.all_models.append(self.cursor.fetchone())
            self.cursor.execute('''SELECT * FROM public.polet WHERE uniq_id_bpla=%s''', (self.all_models[0][0],))
            flight_sam = self.cursor.fetchall()
            time_polets = 0
            dist_polet = 0
            hight_polet = 0
            for row in flight_sam:
                time_polets = time_polets + row[2]
                dist_polet = dist_polet + row[3]
                hight_polet = hight_polet + row[4]

            cv = (self.all_models[0][6] - (time_polets / len(flight_sam))) / self.all_models[0][6]
            cd = (self.all_models[0][7] - (dist_polet / len(flight_sam))) / self.all_models[0][7]
            cb = (self.all_models[0][4] - (hight_polet / len(flight_sam))) / self.all_models[0][4]

            headers = [""] * self.ui_an_jiv.tableWidget.rowCount()
            self.ui_an_jiv.tableWidget.setVerticalHeaderLabels(headers)
            self.index_jivuchesti = ((cv+cd+cb)/3)*100
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][0]))
            self.ui_an_jiv.tableWidget.setItem(0, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][1]))
            self.ui_an_jiv.tableWidget.setItem(0, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][2]))
            self.ui_an_jiv.tableWidget.setItem(0, 2, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.index_jivuchesti))
            self.ui_an_jiv.tableWidget.setItem(0, 3, item)
        else:
            self.all_models=[]
            self.cursor.execute('''SELECT * FROM public.bpla''')
            self.all_models = self.cursor.fetchall()
            self.index_all_models = []
            for row in self.all_models:
                self.cursor.execute('''SELECT * FROM public.polet where uniq_id_bpla=%s''', (row[0],))
                data = self.cursor.fetchall()
                time_polets = 0
                dist_polet = 0
                hight_polet = 0
                for item in data:
                    time_polets = time_polets + item[2]
                    dist_polet = dist_polet + item[3]
                    hight_polet = hight_polet + item[4]
                cv = (row[6] - (time_polets / len(data))) / row[6]
                cd = (row[7] - (dist_polet / len(data))) / row[7]
                cb = (row[4] - (hight_polet / len(data))) / row[4]
                self.index_all_models.append(((cv + cd + cb) / 3) * 100)

            if sort_po == "По возрастанию":
                self.index_all_models.sort()
            elif sort_po == "По убыванию":
                self.index_all_models.sort(reverse=True)

            self.ui_an_jiv.tableWidget.setRowCount(len(self.index_all_models))
            headers = [""]*self.ui_an_jiv.tableWidget.rowCount()
            self.ui_an_jiv.tableWidget.setVerticalHeaderLabels(headers)
            for i, index in enumerate(self.index_all_models):
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][0]))
                self.ui_an_jiv.tableWidget.setItem(i, 0, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][1]))
                self.ui_an_jiv.tableWidget.setItem(i, 1, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][2]))
                self.ui_an_jiv.tableWidget.setItem(i, 2, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.index_all_models[i]))
                self.ui_an_jiv.tableWidget.setItem(i, 3, item)

        self.an_jiv.show()

    def analit_effect(self):
        sort_po = self.ui_fourth.comboBox_SORT_PO.currentText()
        model = self.ui_fourth.comboBox_VIBOR_MODEL.currentText()
        #сначало считать по убыванию ил по возрастанию
        #затем считать по какой модели и вывести
        if model != "Не выбрано":
            self.all_models =[]
            self.cursor.execute('''SELECT * FROM public.bpla WHERE model_bpla=%s''',(model,))
            self.all_models.append(self.cursor.fetchone())
            self.cursor.execute('''SELECT * FROM public.polet WHERE uniq_id_bpla=%s''',(self.all_models[0][0],))
            flight_sam = self.cursor.fetchall()
            complete_task = 0
            time_polets = 0
            dist_polet = 0
            hight_polet = 0
            toplivo_polet = 0
            p = 0
            for row in flight_sam:
                complete_task = complete_task + row[5]
                time_polets = time_polets + row[2]
                dist_polet = dist_polet + row[3]
                hight_polet = hight_polet + row[4]
                toplivo_polet = toplivo_polet + row[7]

            cz = (complete_task - len(flight_sam))/(complete_task*4)
            cv = (self.all_models[0][6] - (time_polets/len(flight_sam)))/self.all_models[0][6]
            cd = (self.all_models[0][7] - (dist_polet/len(flight_sam)))/self.all_models[0][7]
            cb = (self.all_models[0][4] - (hight_polet/len(flight_sam)))/self.all_models[0][4]
            cp = (self.all_models[0][5] - (toplivo_polet/len(flight_sam)))/self.all_models[0][5]
            if self.all_models[0][6] == True:
                p = 0.5
            else:
                p = 1

            headers = [""]*self.ui_an_effe.tableWidget.rowCount()
            self.ui_an_effe.tableWidget.setVerticalHeaderLabels(headers)
            self.index_effective = (cz / cv+cd+cb+cp+1)*p
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][0]))
            self.ui_an_effe.tableWidget.setItem(0, 0, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][1]))
            self.ui_an_effe.tableWidget.setItem(0, 1, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.all_models[0][2]))
            self.ui_an_effe.tableWidget.setItem(0, 2, item)
            item = QtWidgets.QTableWidgetItem()
            item.setText(str(self.index_effective))
            self.ui_an_effe.tableWidget.setItem(0,3,item)
        else:
            self.all_models=[]
            self.cursor.execute("SELECT * FROM bpla")
            self.all_models = self.cursor.fetchall()
            self.index_all_models = []
            for row in self.all_models:
                self.cursor.execute('''SELECT * FROM public.polet where uniq_id_bpla=%s''', (row[0],))
                data = self.cursor.fetchall()
                complete_task = 0
                time_polets = 0
                dist_polet = 0
                hight_polet = 0
                toplivo_polet = 0
                p = 0
                for item in data:
                    complete_task = complete_task + item[5]
                    time_polets = time_polets + item[2]
                    dist_polet = dist_polet + item[3]
                    hight_polet = hight_polet + item[4]
                    toplivo_polet = toplivo_polet + item[7]

                cz = ceil((complete_task - len(data)) / (complete_task * 4+1)) +1
                cv = ceil((row[6] - (time_polets+1 / len(data)+1)) / row[6]) +1
                cd = ceil((row[7] - (dist_polet+1 / len(data)+1)) / row[7]) + 1
                cb = ceil((row[4] - (hight_polet+1 / len(data)+1)) / row[4]) + 1
                cp = ceil((row[5] - (toplivo_polet+1 / len(data)+1)) / row[5]) + 1

                if row[6] == True:
                    p = 0.5
                else:
                    p = 1
                self.index_all_models.append(ceil((cz+1 / cv+cd+cb+cp+1))*p)

            if sort_po == "По возрастанию":
                self.index_all_models.sort()
            elif sort_po == "По убыванию":
                self.index_all_models.sort(reverse=True)

            self.ui_an_effe.tableWidget.setRowCount(len(self.index_all_models))
            headers = [""]*self.ui_an_effe.tableWidget.rowCount()
            self.ui_an_effe.tableWidget.setVerticalHeaderLabels(headers)
            for i, index in enumerate(self.index_all_models):
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][0]))
                self.ui_an_effe.tableWidget.setItem(i, 0, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][1]))
                self.ui_an_effe.tableWidget.setItem(i, 1, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.all_models[i][2]))
                self.ui_an_effe.tableWidget.setItem(i, 2, item)
                item = QtWidgets.QTableWidgetItem()
                item.setText(str(self.index_all_models[i]))
                self.ui_an_effe.tableWidget.setItem(i, 3, item)

        self.effect_widget.show()


    def can_vivod(self):
        stroka = self.ui_an_effe.tableWidget.selectionModel().selectedRows()
        if stroka:
            row = stroka[0].row()
            self.vivod_v_fail_ef_bpla = self.ui_an_effe.tableWidget.item(row,0).text()
            self.ui_an_effe.menu.setEnabled(True)
        else:
            self.ui_an_effe.menu.setEnabled(False)

    def can_vivo2(self):
        stroka = self.ui_an_jiv.tableWidget.selectionModel().selectedRows()
        if stroka:
            row = stroka[0].row()
            self.vivod_v_fail_jiv_bpla = self.ui_an_jiv.tableWidget.item(row,0).text()
            self.ui_an_jiv.menu.setEnabled(True)
        else:
            self.ui_an_jiv.menu.setEnabled(False)

    def can_vivod3(self):
        stroka = self.ui_an_res.tableWidget.selectionModel().selectedRows()
        if stroka:
            row = stroka[0].row()
            self.vivod_v_fail_res_bpla = self.ui_an_res.tableWidget.item(row,0).text()
            self.ui_an_res.menu.setEnabled(True)
        else:
            self.ui_an_res.menu.setEnabled(False)

    def vivod_v_fail_effect(self):
        stroka = self.ui_an_effe.tableWidget.selectionModel().selectedRows()
        data_rows = []
        for index in stroka:
            row_item = []
            for column in range(self.ui_an_effe.tableWidget.columnCount()):
                item = self.ui_an_effe.tableWidget.item(index.row(),column)
                row_item.append(item.text())
            data_rows.append(row_item)

        data_colums = []
        for column in range(self.ui_an_effe.tableWidget.columnCount()):
            item = self.ui_an_effe.tableWidget.horizontalHeaderItem(column)
            data_colums.append(item.text())

        vac = dict()
        for i in range(len(data_colums)):
            vac[data_colums[i]] = []
            for j in range(len(data_rows)):
                if i == 0:
                    vac[data_colums[i]].append(int(data_rows[j][i]))
                elif i == 3:
                    vac[data_colums[i]].append(float(data_rows[j][i]))
                else:
                    vac[data_colums[i]] += [data_rows[j][i]]

        bottoms = Side(border_style='medium',color='000000')
        table = pd.DataFrame(vac)
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        for i,row in enumerate(dataframe_to_rows(table,index=False),1):
            ws.append(row)
            for j,cell in enumerate(ws[i]):
                if i == 1:
                    letter = chr(65+j)
                    ws.column_dimensions[letter].width = 25
                cell.border = Border(left = bottoms,right = bottoms,top = bottoms,bottom=bottoms)

        ws2 = wb.create_sheet("Диаграмма")

        val = Reference(ws,min_col=4,max_col=4,min_row=2,max_row=7)

        c1 = BarChart()
        c1.add_data(val)
        c1.title = "Стобчатая диаграмма"
        c1.width = 15
        c1.height = 11
        ws2.add_chart(c1,'M5')

        c2 = PieChart3D()
        c2.add_data(val)
        c2.title = "Круговая диаграмма"
        c2.width = 15
        c2.height = 11
        ws2.add_chart(c2,'C5')

        wb.save("./report.xlsx")


    def vivod_v_fail_res(self):
        stroka = self.ui_an_res.tableWidget.selectionModel().selectedRows()
        data_rows = []
        for index in stroka:
            row_item = []
            for column in range(self.ui_an_res.tableWidget.columnCount()):
                item = self.ui_an_res.tableWidget.item(index.row(),column)
                if column == 0:
                    row_item.append(int(item.text()))
                elif column == 3:
                    row_item.append(float(item.text()))
                else:
                    row_item.append(item.text())
            data_rows.append(row_item)

        data_column = []
        for column in range(self.ui_an_res.tableWidget.columnCount()):
            item = self.ui_an_res.tableWidget.horizontalHeaderItem(column)
            data_column.append(item.text())
        vac = dict()
        for i in range(len(data_column)):
            vac[data_column[i]] = []
            for j in range(len(data_rows)):
                vac[data_column[i]] += [data_rows[j][i]]

        bottoms = Side(border_style="medium",color="000000")
        table = pd.DataFrame(vac)
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        for i,row in enumerate(dataframe_to_rows(table,index=False),1):
            ws.append(row)
            for j,cell in enumerate(ws[i]):
                if i == 1:
                    letter = chr(65+j)
                    ws.column_dimensions[letter].width = 25
                cell.border = Border(left = bottoms,right = bottoms,top = bottoms,bottom=bottoms)
        # bar_chart
        # pie_chart
        # __3d

        ws2 = wb.create_sheet("Диаграммы")

        v1 = Reference(ws,min_col=4,min_row=2,max_row=len(data_rows)+1,max_col=4)

        c1 = BarChart()
        c1.add_data(v1)
        c1.title = "Стобчатая диаграмма"
        c1.width = 15
        c1.height = 11
        c1.legend.legendPos = 'b'
        ws2.add_chart(c1,"M5")

        c2 = PieChart3D()
        c2.add_data(v1)
        c2.width = 15
        c2.height = 11
        c2.legend.legendPos = 'b'
        c2.title = "Круговая диаграмма"
        ws2.add_chart(c2,"C5")

        wb.save("./report.xlsx")







    def vivod_v_fail_jiv(self):
        stroka = self.ui_an_jiv.tableWidget.selectionModel().selectedRows()
        data_rows = []
        for index in stroka:
            row_item = []
            for column in range(self.ui_an_jiv.tableWidget.columnCount()):
                item = self.ui_an_jiv.tableWidget.item(index.row(),column)
                row_item.append(item.text())
            data_rows.append(row_item)
#ПРЕДУСМОТРЕТЬ ПРОВЕРКУ КОГДА ПОЛЕТОВ У БПЛА НЕТ
        data_colums = []
        for column in range(self.ui_an_jiv.tableWidget.columnCount()):
                item = self.ui_an_jiv.tableWidget.horizontalHeaderItem(column)
                data_colums.append(item.text())
        vac = dict()
        for i in range(len(data_colums)):
            vac[data_colums[i]] = []
            for j in range(len(data_rows)):
                if i == 0:
                    vac[data_colums[i]].append(int(data_rows[j][i]))
                elif i == 3:
                    vac[data_colums[i]].append(float(data_rows[j][i]))
                else:
                    vac[data_colums[i]] += [data_rows[j][i]]

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"

        table = pd.DataFrame(vac)
        bottoms = Side(border_style='medium',color="000000")
        for i,row in enumerate(dataframe_to_rows(table,index=False),1):
            ws.append(row)
            for j,cell in enumerate(ws[i]):
                if i == 1:
                    letter = chr(65+j)
                    ws.column_dimensions[letter].width = 25
                cell.border = Border(left = bottoms,right = bottoms,top = bottoms,bottom=bottoms)

        ws2 = wb.create_sheet("Диаграммы")

        val = Reference(ws,min_col=4,max_col=4,min_row=2,max_row=len(data_rows)+1)

        с1 = BarChart()
        с1.add_data(val)
        с1.title = "Столбчатая диаграмма"
        с1.legend.position = 'b'
        с1.width = 15
        с1.height = 11
        ws2.add_chart(с1,"M5")

        chart = PieChart3D()
        chart.add_data(val)
        chart.title = "Круговая диаграмма"
        chart.legend.position = 'b'
        chart.width = 15
        chart.height = 11
        ws2.add_chart(chart,"C5")

        wb.save('./report.xlsx')

    def connect_to_bd(self):
        try:
            self.connect = pg.connect(
                host="localhost",
                user="postgres",
                password="Roma23082003",
                database="football_club",
                port=5432
            )
            self.cursor = self.connect.cursor()
            print("Connect estblished")
        except Exception as e:
            print(f"Error connecting {e}")

